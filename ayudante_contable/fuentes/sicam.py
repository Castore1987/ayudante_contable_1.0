"""Lector de SICAM (autónomos y monotributo): situación de revista y deuda.

SICAM exporta a PDF **sin capa de texto**: los caracteres vienen convertidos a
contornos vectoriales, así que no hay nada que extraer con un parser de texto.
La única vía es OCR.

Para que el OCR no dependa de adivinar dónde empieza cada columna, se aprovecha
que el PDF **sí** conserva los bordes de la tabla como vectores. De ahí se
deriva la grilla exacta (columnas y filas) y cada palabra reconocida se asigna
a su celda por posición. Es mucho más confiable que partir el texto por
espacios: las celdas que envuelven en varias líneas —«Art. 1 Ley 25321» ocupa
tres— se reconstruyen solas.

Dos documentos, dos propósitos:

``Situación de Revista``
    Los períodos de actividad como autónomo/monotributista, con su código de
    actividad y categoría. Es la obligación: de acá sale qué meses debía
    aportar. Trae además una columna de ``Beneficios`` donde aparece el
    Art. 1 Ley 25.321.

``Detalle de la Deuda``
    Mes a mes, cuántos meses se adeudan y por cuánto. Un período con
    ``Meses Adeudados`` en cero y total en cero está cancelado. La columna
    ``Benef. Aplic.`` marca los períodos alcanzados por el Art. 1 Ley 25.321.

Lo que el documento **no** dice: si una deuda está en plan de pagos o en
moratoria. Eso no figura en el archivo y lo aporta el estudio como política
(ver ``PoliticaDeuda``).
"""

from __future__ import annotations

import re
import shutil
from datetime import datetime
import subprocess
import tempfile
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Sequence

from ..modelo.dominio import EstadoIngreso, Periodo, TipoAporte, normalizar_cuil
from .base import ErrorFuente
from .planilla import parsear_decimal

__all__ = [
    "leer_revista",
    "leer_deuda",
    "leer_revista_excel",
    "leer_deuda_excel",
    "es_excel",
    "PeriodoRevista",
    "FilaDeuda",
    "LecturaRevista",
    "LecturaDeuda",
    "es_sicam",
    "leer_sicam",
    "historia_desde_sicam",
    "PoliticaDeuda",
    "ART1_LEY_25321",
    "OCR_DPI",
]

ART1_LEY_25321 = "art. 1 ley 25.321"
OCR_DPI = 300
CERO = Decimal("0")

# Códigos de actividad de la situación de revista que NO generan obligación de
# aportar. Un tramo con uno de estos códigos declara actividad, pero sus meses
# no suman al cómputo de aportes: no hay aporte que exigir ni que acreditar.
CODIGOS_NO_APORTANTES = frozenset({"11"})

# El OCR arrastra los bordes de la tabla como "|", "[", "]" y a veces parte el
# año ("09/1 992"). Todo eso se limpia antes de interpretar la celda.
# Bordes y filetes de la tabla que el OCR devuelve como caracteres.
_RE_BORDES = re.compile(r"[|\[\]{}¦<>—–_·¡]+")
_RE_PERIODO = re.compile(r"(\d{1,2})\s*/\s*(\d{4})")
# El OCR a veces parte el año ("09/1 992"): se cierran los espacios entre dígitos.
_RE_ESPACIO_ENTRE_DIGITOS = re.compile(r"(?<=\d)\s+(?=\d)")
_RE_LEY_25321 = re.compile(r"ley\d*25321|art\d*ley")
_RE_NO_ALFANUM = re.compile(r"[^a-z0-9]+")
_RE_ESPACIOS = re.compile(r"\s+")


# --------------------------------------------------------------- herramientas


def _verificar_ocr() -> str:
    ruta = shutil.which("tesseract")
    if ruta is None:
        raise ErrorFuente(
            "Los PDF de SICAM no tienen capa de texto: se necesitan OCR y un "
            "renderizador.\n"
            "    apt-get install -y tesseract-ocr tesseract-ocr-spa\n"
            "    pip install pymupdf\n"
            "Alternativa: transcribir la situación de revista y la deuda a una "
            "planilla CSV y usarla como fuente."
        )
    return ruta


def _abrir_pdf(ruta: Path):
    try:
        import pymupdf
    except ImportError as exc:
        raise ErrorFuente(
            "Renderizar los PDF de SICAM necesita pymupdf:  pip install pymupdf"
        ) from exc
    return pymupdf.open(str(ruta))


def _limites(valores: Sequence[float], tolerancia: float = 2.0) -> list[float]:
    """Agrupa coordenadas cercanas en un solo límite."""
    agrupados: list[float] = []
    for valor in sorted(valores):
        if not agrupados or valor - agrupados[-1] > tolerancia:
            agrupados.append(valor)
    return agrupados


def _grilla(pagina, minimo_repeticiones: int = 20) -> tuple[list[tuple[float, float]], list[float]]:
    """Deriva columnas y límites de fila desde los bordes vectoriales.

    Devuelve ``([(x0, x1), …], [y0, y1, …])``. Las columnas se detectan por los
    segmentos horizontales que se repiten en casi todas las filas: su ``x0`` y
    su ancho definen la celda.
    """
    segmentos = [
        r for r in pagina.rects if r["height"] <= 2 and 8 < r["width"] < 260
    ]
    if not segmentos:
        raise ErrorFuente(
            "No encontré la grilla de la tabla en el PDF de SICAM. "
            "¿Es el documento correcto?"
        )

    conteo: dict[int, int] = {}
    anchos: dict[int, dict[int, int]] = {}
    for segmento in segmentos:
        x = round(segmento["x0"])
        conteo[x] = conteo.get(x, 0) + 1
        ancho = round(segmento["width"])
        por_ancho = anchos.setdefault(x, {})
        por_ancho[ancho] = por_ancho.get(ancho, 0) + 1

    columnas: list[tuple[float, float]] = []
    for x in sorted(conteo):
        if conteo[x] < minimo_repeticiones:
            continue
        ancho = max(anchos[x].items(), key=lambda par: par[1])[0]
        columnas.append((float(x), float(x + ancho)))

    if not columnas:
        raise ErrorFuente("No pude determinar las columnas de la tabla de SICAM.")

    filas = _limites([s["top"] for s in segmentos])
    return columnas, filas


def _ocr_fila(
    pagina_pdf, y0: float, y1: float, ancho: float, dpi: int
) -> list[tuple[float, str]]:
    """OCR de una franja horizontal (un renglón); devuelve ``(x_en_puntos, texto)``.

    Se recorta renglón por renglón y no columna por columna. Recortar columnas
    parecía más directo, pero una franja de 30 puntos de ancho corta los glifos
    de los bordes y el OCR devuelve "0519" donde dice "05/1995". Un renglón
    completo, en cambio, es una sola línea de texto ancha: tesseract no fusiona
    columnas —las separa el espacio en blanco— y ningún caracter queda cortado.
    """
    import pymupdf

    _verificar_ocr()
    escala = dpi / 72.0
    margen = 2.0
    recorte = pymupdf.Rect(0, max(y0 - margen, 0), ancho, y1 + margen)
    pixmap = pagina_pdf.get_pixmap(dpi=dpi, clip=recorte)

    with tempfile.TemporaryDirectory() as carpeta:
        imagen = Path(carpeta) / "fila.png"
        pixmap.save(str(imagen))
        salida = subprocess.run(
            ["tesseract", str(imagen), "stdout", "-l", "spa", "--psm", "6", "tsv"],
            capture_output=True,
            text=True,
            timeout=120,
        )
    if salida.returncode != 0:
        raise ErrorFuente(f"El OCR falló: {salida.stderr.strip()[:200]}")

    palabras: list[tuple[float, str]] = []
    for linea in salida.stdout.splitlines()[1:]:
        campos = linea.split("\t")
        if len(campos) < 12:
            continue
        texto = campos[11].strip()
        if not texto or texto == "-1":
            continue
        try:
            izquierda, ancho_caja = int(campos[6]), int(campos[8])
        except ValueError:
            continue
        palabras.append(((izquierda + ancho_caja / 2) / escala, texto))
    return palabras


def _ocr_pagina(pagina_pdf, dpi: int) -> list[tuple[float, float, str]]:
    """OCR de la página entera; devuelve ``(centro_x, centro_y, texto)``."""
    _verificar_ocr()
    escala = dpi / 72.0
    pixmap = pagina_pdf.get_pixmap(dpi=dpi)

    with tempfile.TemporaryDirectory() as carpeta:
        imagen = Path(carpeta) / "pagina.png"
        pixmap.save(str(imagen))
        salida = subprocess.run(
            ["tesseract", str(imagen), "stdout", "-l", "spa", "--psm", "6", "tsv"],
            capture_output=True,
            text=True,
            timeout=180,
        )
    if salida.returncode != 0:
        raise ErrorFuente(f"El OCR falló: {salida.stderr.strip()[:200]}")

    palabras: list[tuple[float, float, str]] = []
    for linea in salida.stdout.splitlines()[1:]:
        campos = linea.split("\t")
        if len(campos) < 12:
            continue
        texto = campos[11].strip()
        if not texto or texto == "-1":
            continue
        try:
            izq, arriba, ancho_caja, alto_caja = (int(campos[i]) for i in (6, 7, 8, 9))
        except ValueError:
            continue
        palabras.append(
            (
                (izq + ancho_caja / 2) / escala,
                (arriba + alto_caja / 2) / escala,
                texto,
            )
        )
    return palabras


def _celdas_pagina(
    pagina_pdf,
    columnas: list[tuple[float, float]],
    filas: list[float],
    dpi: int,
) -> dict[tuple[int, int], str]:
    """Celdas a partir de un solo OCR de la página completa.

    Sirve para tablas de pocas columnas anchas, como la situación de revista:
    ahí el OCR de página entera reconoce todos los renglones. En la tabla de
    deuda, con diecisiete columnas angostas, fusiona celdas y hay que ir
    renglón por renglón (:func:`_celdas`).
    """
    contenido: dict[tuple[int, int], list[tuple[float, float, str]]] = {}

    for centro_x, centro_y, texto in _ocr_pagina(pagina_pdf, dpi):
        columna = next(
            (i for i, (x0, x1) in enumerate(columnas) if x0 - 3 <= centro_x <= x1 + 3),
            None,
        )
        if columna is None:
            continue
        fila = next(
            (i for i in range(len(filas) - 1) if filas[i] - 1 <= centro_y < filas[i + 1] - 1),
            None,
        )
        if fila is None:
            continue
        contenido.setdefault((fila, columna), []).append((centro_y, centro_x, texto))

    return {
        clave: _RE_ESPACIOS.sub(" ", " ".join(t for *_, t in sorted(partes))).strip()
        for clave, partes in contenido.items()
    }


def _celdas(
    pagina_pdf,
    columnas: list[tuple[float, float]],
    filas: list[float],
    dpi: int,
    ancho: float,
) -> dict[tuple[int, int], str]:
    """Arma el contenido de cada celda haciendo OCR renglón por renglón."""
    contenido: dict[tuple[int, int], list[tuple[float, str]]] = {}

    for fila in range(len(filas) - 1):
        alto = filas[fila + 1] - filas[fila]
        if alto < 4:            # dos bordes pegados, no es un renglón
            continue
        for centro_x, texto in _ocr_fila(
            pagina_pdf, filas[fila], filas[fila + 1], ancho, dpi
        ):
            columna = next(
                (i for i, (x0, x1) in enumerate(columnas) if x0 - 3 <= centro_x <= x1 + 3),
                None,
            )
            if columna is None:
                continue
            contenido.setdefault((fila, columna), []).append((centro_x, texto))

    return {
        clave: _RE_ESPACIOS.sub(" ", " ".join(t for _, t in sorted(partes))).strip()
        for clave, partes in contenido.items()
    }


def _limpiar(texto: str) -> str:
    """Quita los bordes de tabla que el OCR mete dentro de la celda."""
    sin_bordes = _RE_BORDES.sub(" ", texto)
    return _RE_ESPACIO_ENTRE_DIGITOS.sub("", _RE_ESPACIOS.sub(" ", sin_bordes)).strip()


def _texto(celdas: dict, fila: int, columna: int) -> str:
    return _limpiar(celdas.get((fila, columna), ""))


def _periodo(texto: str) -> Periodo | None:
    coincidencia = _RE_PERIODO.search(_limpiar(texto))
    if not coincidencia:
        return None
    mes = int(coincidencia.group(1))
    anio = int(coincidencia.group(2))
    if not (1 <= mes <= 12 and 1900 <= anio <= 2200):
        return None
    return Periodo(anio, mes)


def _tiene_art1(texto: str) -> bool:
    """Reconoce «Art. 1 Ley 25321» pese al ruido del OCR."""
    return bool(_RE_LEY_25321.search(_RE_NO_ALFANUM.sub("", texto.lower())))


# ------------------------------------------------------- situación de revista


@dataclass(frozen=True)
class PeriodoRevista:
    """Un tramo de actividad declarado en SICAM."""

    inicio: Periodo
    cese: Periodo | None
    codigo_actividad: str = ""
    tipo_sociedad: str = ""
    categoria_optativa: str = ""
    beneficio_desde: Periodo | None = None
    beneficio_hasta: Periodo | None = None
    tipo_beneficio: str = ""

    @property
    def prescripto_art1(self) -> bool:
        return _tiene_art1(self.tipo_beneficio)

    def meses_beneficio(self) -> set[Periodo]:
        if self.beneficio_desde is None or self.beneficio_hasta is None:
            return set()
        return set(Periodo.rango(self.beneficio_desde, self.beneficio_hasta))


@dataclass
class LecturaRevista:
    cuil: str | None = None
    periodos: list[PeriodoRevista] = field(default_factory=list)
    advertencias: list[str] = field(default_factory=list)

    def meses_prescriptos(self) -> set[Periodo]:
        """Meses alcanzados por el Art. 1 Ley 25.321 según la revista."""
        meses: set[Periodo] = set()
        for periodo in self.periodos:
            if periodo.prescripto_art1:
                meses |= periodo.meses_beneficio()
        return meses


def leer_revista(ruta: str | Path, dpi: int = OCR_DPI) -> LecturaRevista:
    """Lee la «Situación de Revista» de SICAM, en planilla o en PDF."""
    if es_excel(ruta):
        return leer_revista_excel(ruta)
    ruta = Path(ruta)
    if not ruta.exists():
        raise ErrorFuente(f"No encontré el archivo: {ruta}")

    import pdfplumber

    lectura = LecturaRevista()
    documento = _abrir_pdf(ruta)
    try:
        with pdfplumber.open(str(ruta)) as pdf:
            for numero, pagina in enumerate(pdf.pages):
                columnas, filas = _grilla(pagina, minimo_repeticiones=4)
                celdas = _celdas_pagina(documento[numero], columnas, filas, dpi)
                _filas_revista(celdas, len(filas), lectura)
    finally:
        documento.close()

    if not lectura.periodos:
        raise ErrorFuente(
            f"No reconocí ningún período en {ruta.name}. Si el OCR falló, "
            "revisá la calidad del PDF o transcribí la tabla a CSV."
        )
    return lectura


def _filas_revista(celdas: dict, cantidad_filas: int, lectura: LecturaRevista) -> None:
    for fila in range(cantidad_filas):
        inicio = _periodo(_texto(celdas, fila, 0))
        if inicio is None:
            continue
        lectura.periodos.append(
            PeriodoRevista(
                inicio=inicio,
                cese=_periodo(_texto(celdas, fila, 1)),
                codigo_actividad=_texto(celdas, fila, 2),
                tipo_sociedad=_texto(celdas, fila, 3),
                categoria_optativa=_texto(celdas, fila, 4),
                beneficio_desde=_periodo(_texto(celdas, fila, 9)),
                beneficio_hasta=_periodo(_texto(celdas, fila, 10)),
                tipo_beneficio=_texto(celdas, fila, 11),
            )
        )


# ------------------------------------------------------------ detalle de deuda


@dataclass(frozen=True)
class FilaDeuda:
    """Un renglón del ``Detalle de la Deuda``."""

    desde: Periodo
    hasta: Periodo
    categoria_historica: str = ""
    categoria_actual: str = ""
    beneficio_aplicado: str = ""
    meses_aportes: Decimal = CERO
    capital_subtotal: Decimal = CERO
    intereses_subtotal: Decimal = CERO
    total: Decimal = CERO

    @property
    def prescripto_art1(self) -> bool:
        return _tiene_art1(self.beneficio_aplicado)

    @property
    def adeuda(self) -> bool:
        """Hay deuda real.

        Se decide por el total y no por los meses adeudados: la columna de
        meses queda pegada al texto del beneficio y el OCR la lee peor, mientras
        que el importe total sale limpio. Un renglón en cero no tiene nada que
        reclamar, cualquiera sea el motivo.
        """
        return self.total > CERO

    @property
    def cancelado(self) -> bool:
        return self.total <= CERO

    @property
    def coherente(self) -> bool:
        """¿El total coincide con capital + intereses?

        Es el control de calidad del OCR: si los tres importes cierran entre
        sí, se leyeron bien. Si no cierran, el renglón se informa como dudoso
        en lugar de darlo por bueno.
        """
        esperado = self.capital_subtotal + self.intereses_subtotal
        if esperado == CERO and self.total == CERO:
            return True
        if esperado == CERO:
            return False
        return abs(self.total - esperado) <= max(esperado * Decimal("0.001"), Decimal("1"))

    def meses(self) -> set[Periodo]:
        return set(Periodo.rango(self.desde, self.hasta))


@dataclass
class LecturaDeuda:
    cuil: str | None = None
    filas: list[FilaDeuda] = field(default_factory=list)
    advertencias: list[str] = field(default_factory=list)

    @property
    def total_adeudado(self) -> Decimal:
        return sum((f.total for f in self.filas), start=CERO)

    def meses_prescriptos(self) -> set[Periodo]:
        meses: set[Periodo] = set()
        for fila in self.filas:
            if fila.prescripto_art1:
                meses |= fila.meses()
        return meses

    def meses_con_deuda(self) -> set[Periodo]:
        meses: set[Periodo] = set()
        for fila in self.filas:
            if fila.adeuda and not fila.prescripto_art1:
                meses |= fila.meses()
        return meses

    def meses_cancelados(self) -> set[Periodo]:
        meses: set[Periodo] = set()
        for fila in self.filas:
            if fila.cancelado and not fila.prescripto_art1:
                meses |= fila.meses()
        return meses


def leer_deuda(ruta: str | Path, dpi: int = OCR_DPI) -> LecturaDeuda:
    """Lee el «Detalle de la Deuda» de SICAM, en planilla o en PDF."""
    if es_excel(ruta):
        return leer_deuda_excel(ruta)
    ruta = Path(ruta)
    if not ruta.exists():
        raise ErrorFuente(f"No encontré el archivo: {ruta}")

    import pdfplumber

    lectura = LecturaDeuda()
    documento = _abrir_pdf(ruta)
    try:
        with pdfplumber.open(str(ruta)) as pdf:
            for numero, pagina in enumerate(pdf.pages):
                try:
                    columnas, filas = _grilla(pagina)
                except ErrorFuente:
                    continue          # páginas de botones, sin tabla
                if len(columnas) < 15:
                    continue
                celdas = _celdas(
                    documento[numero], columnas, filas, dpi, float(pagina.width)
                )
                _filas_deuda(celdas, len(filas), lectura, numero + 1)
    finally:
        documento.close()

    if not lectura.filas:
        raise ErrorFuente(
            f"No reconocí ningún renglón de deuda en {ruta.name}. Si el OCR falló, "
            "revisá la calidad del PDF."
        )
    return lectura


def _periodos_del_renglon(celdas: dict, fila: int) -> tuple[Periodo | None, Periodo | None]:
    """Lee Desde y Hasta juntos.

    El borde entre ambas columnas a veces se reconoce como un caracter y las
    fusiona en un solo token ("07/1980/[09/1980"), así que se buscan todos los
    períodos del par de celdas y se toman en orden.
    """
    texto = f"{_texto(celdas, fila, 0)} {_texto(celdas, fila, 1)}"
    encontrados: list[Periodo] = []
    for coincidencia in _RE_PERIODO.finditer(_limpiar(texto)):
        mes = int(coincidencia.group(1))
        anio = int(coincidencia.group(2))
        if 1 <= mes <= 12 and 1900 <= anio <= 2200:
            encontrados.append(Periodo(anio, mes))
    if not encontrados:
        return None, None
    if len(encontrados) == 1:
        return encontrados[0], encontrados[0]
    return encontrados[0], encontrados[1]


def _filas_deuda(
    celdas: dict, cantidad_filas: int, lectura: LecturaDeuda, pagina: int
) -> None:
    reconocidas = 0
    candidatas = 0
    for fila in range(cantidad_filas):
        crudo = f"{_texto(celdas, fila, 0)}{_texto(celdas, fila, 1)}"
        if any(c.isdigit() for c in crudo):
            candidatas += 1
        desde, hasta = _periodos_del_renglon(celdas, fila)
        if desde is None or hasta is None:
            continue
        reconocidas += 1
        if hasta < desde:
            lectura.advertencias.append(
                f"Renglón de deuda con período invertido ({desde} a {hasta}); se omitió."
            )
            continue
        # El texto de "Benef. Aplic." envuelve en tres líneas y se derrama a
        # las celdas vecinas, así que se busca en las tres columnas.
        beneficio = " ".join(
            _texto(celdas, fila, c) for c in (2, 3, 4)
        ).strip()

        lectura.filas.append(
            FilaDeuda(
                desde=desde,
                hasta=hasta,
                categoria_historica=_texto(celdas, fila, 2),
                categoria_actual=_texto(celdas, fila, 3),
                beneficio_aplicado=beneficio,
                meses_aportes=parsear_decimal(_texto(celdas, fila, 5)) or CERO,
                capital_subtotal=parsear_decimal(_texto(celdas, fila, 11)) or CERO,
                intereses_subtotal=parsear_decimal(_texto(celdas, fila, 15)) or CERO,
                total=parsear_decimal(_texto(celdas, fila, 16)) or CERO,
            )
        )

    incoherentes = [f for f in lectura.filas if not f.coherente]
    if incoherentes:
        lectura.advertencias.append(
            f"{len(incoherentes)} renglón(es) de deuda no cierran capital + "
            "intereses = total: el OCR pudo leer mal un importe. Están marcados "
            "como dudosos y hay que cotejarlos contra el PDF."
        )

    if candidatas and reconocidas < candidatas:
        lectura.advertencias.append(
            f"Página {pagina} del detalle de deuda: se reconocieron {reconocidas} "
            f"de {candidatas} renglones con datos. Los no reconocidos quedaron "
            "fuera del informe; conviene cotejar esa página contra el PDF."
        )


def es_sicam(ruta: str | Path) -> bool:
    """¿El PDF viene de SICAM? Se reconoce por el encabezado impreso."""
    try:
        documento = _abrir_pdf(Path(ruta))
    except ErrorFuente:
        return False
    try:
        if len(documento) == 0:
            return False
        palabras = _ocr_palabras(documento[0], dpi=150)
    except ErrorFuente:
        return False
    finally:
        documento.close()
    texto = " ".join(t for *_, t in palabras).upper()
    return "SICAM" in texto


# ---------------------------------------------- de SICAM a historia laboral


@dataclass(frozen=True)
class PoliticaDeuda:
    """Cómo se interpreta una deuda de autónomos que figura en SICAM.

    El documento informa la deuda pero **no** dice si está incluida en un plan
    de pagos o en una moratoria: ese dato lo aporta el estudio. Por eso es una
    política explícita y no una deducción del archivo.

    ``deuda_es_regularizada`` en ``True`` es el criterio indicado por el
    estudio: la deuda que aparece está en plan de pagos o moratoria, así que el
    mes computa como aportado y queda señalado. En ``False`` la deuda se trata
    como aporte no ingresado y el mes no computa.
    """

    deuda_es_regularizada: bool = True
    # El Art. 1 Ley 25.321 prescribe el período: no computa ni se reclama.
    aplicar_prescripcion_art1: bool = True

    @property
    def descripcion(self) -> str:
        deuda = (
            "la deuda de SICAM se considera regularizada (plan de pagos o moratoria)"
            if self.deuda_es_regularizada
            else "la deuda de SICAM se considera aporte no ingresado"
        )
        art1 = (
            "los períodos con Art. 1 Ley 25.321 se excluyen por prescripción"
            if self.aplicar_prescripcion_art1
            else "los períodos con Art. 1 Ley 25.321 se computan igual"
        )
        return f"{deuda}; {art1}"


def historia_desde_sicam(
    revista: LecturaRevista,
    deuda: LecturaDeuda | None,
    cuil: str,
    nombre: str | None = None,
    politica: PoliticaDeuda | None = None,
    fuente: str = "sicam",
):
    """Arma la historia laboral de autónomos/monotributo a partir de SICAM.

    La situación de revista dice **qué meses había obligación de aportar**; el
    detalle de deuda dice **cuáles quedaron impagos**. El cruce de ambos es lo
    que permite afirmar que un mes de autónomo está aportado.
    """
    from ..modelo.dominio import HistoriaLaboral, RegistroMensual

    politica = politica or PoliticaDeuda()
    deuda = deuda or LecturaDeuda()

    prescriptos = revista.meses_prescriptos() | deuda.meses_prescriptos()
    con_deuda = deuda.meses_con_deuda()
    cancelados = deuda.meses_cancelados()
    informa_deuda = bool(deuda.filas)

    # Etiqueta del tramo, por período, tomada de la situación de revista.
    # Los tramos con código no aportante se saltean: declaran actividad, pero
    # no generan meses computables. Si un mes de esos igual tiene renglón en el
    # detalle de deuda, entra más abajo por esa vía, que sí prueba obligación.
    etiqueta: dict[Periodo, str] = {}
    meses_no_aportantes: set[Periodo] = set()
    for periodo in revista.periodos:
        if periodo.cese is None:
            continue
        if periodo.codigo_actividad.strip() in CODIGOS_NO_APORTANTES:
            meses_no_aportantes |= set(Periodo.rango(periodo.inicio, periodo.cese))
            continue
        rotulo = f"Autónomo (act. {periodo.codigo_actividad or 's/d'}"
        rotulo += f", cat. {periodo.categoria_optativa})" if periodo.categoria_optativa else ")"
        for mes in Periodo.rango(periodo.inicio, periodo.cese):
            etiqueta[mes] = rotulo

    # Un mes con deuda registrada tuvo obligación, aunque la revista no lo cubra.
    for mes in con_deuda | cancelados | prescriptos:
        etiqueta.setdefault(mes, "Autónomo (según detalle de deuda)")

    registros = []
    for mes in sorted(etiqueta):
        if mes in prescriptos and politica.aplicar_prescripcion_art1:
            estado = EstadoIngreso.PRESCRIPTO
            nota = "Art. 1 Ley 25.321: período prescripto, no se contabiliza"
        elif mes in con_deuda:
            estado = (
                EstadoIngreso.REGULARIZADO
                if politica.deuda_es_regularizada
                else EstadoIngreso.NO_INGRESADO
            )
            nota = (
                "deuda en SICAM, considerada regularizada por plan de pagos o moratoria"
                if politica.deuda_es_regularizada
                else "deuda en SICAM sin regularizar"
            )
        elif mes in cancelados:
            estado = EstadoIngreso.INGRESADO
            nota = "sin deuda en SICAM"
        else:
            estado = EstadoIngreso.DESCONOCIDO
            nota = (
                "sin renglón de deuda que cubra el período"
                if informa_deuda
                else "SICAM sin detalle de deuda"
            )

        registros.append(
            RegistroMensual(
                periodo=mes,
                cuit_empleador=None,
                empleador=etiqueta[mes],
                tipo=TipoAporte.AUTONOMO,
                remuneracion_imponible=CERO,
                servicio_reconocido=True,
                estado_ingreso=estado,
                observaciones=nota,
            )
        )

    advertencias = list(revista.advertencias) + list(deuda.advertencias)
    advertencias.append(f"Política aplicada a SICAM: {politica.descripcion}.")

    solo_no_aportantes = meses_no_aportantes - set(etiqueta)
    if solo_no_aportantes:
        codigos = ", ".join(sorted(CODIGOS_NO_APORTANTES))
        advertencias.append(
            f"{len(solo_no_aportantes)} mes(es) declarados en la revista bajo un "
            f"código de actividad no aportante ({codigos}) quedaron fuera del "
            "cómputo: declaran actividad, pero no generan aporte."
        )
    if not informa_deuda:
        advertencias.append(
            "SICAM llegó sin detalle de deuda: los meses de autónomo quedan con "
            "ingreso «sin dato»."
        )
    dudosos = [f for f in deuda.filas if not f.coherente]
    if dudosos:
        advertencias.append(
            f"{len(dudosos)} renglón(es) de deuda quedaron marcados como dudosos por "
            "el control de coherencia del OCR; cotejalos contra el PDF antes de firmar."
        )

    historia = HistoriaLaboral(
        cuil=cuil,
        registros=registros,
        nombre=nombre,
        fuente=fuente,
        advertencias_origen=advertencias,
    )
    historia.lectura_revista = revista
    historia.lectura_deuda = deuda
    return historia


# ----------------------------------------------------------------- vía Excel
#
# SICAM también exporta los dos reportes en planilla. Es el camino preferible:
# los importes y los períodos vienen exactos, sin OCR de por medio, y se leen
# los renglones que el PDF vectorizado dejaba fuera de alcance.

_EXTENSIONES_EXCEL = {".xlsx", ".xlsm", ".xltx"}

# Columnas del «Detalle de la Deuda» exportado.
_XL_DEUDA_DESDE, _XL_DEUDA_HASTA = 0, 1
_XL_DEUDA_CAT_HISTORICA, _XL_DEUDA_CAT_ACTUAL = 2, 3
_XL_DEUDA_BENEFICIO = 4
_XL_DEUDA_MESES_APORTES = 5
_XL_DEUDA_CAPITAL_SUBTOTAL = 11
_XL_DEUDA_INTERESES_SUBTOTAL = 15
_XL_DEUDA_TOTAL = 16

# Columnas de la «Situación de Revista» exportada.
_XL_REV_INICIO, _XL_REV_CESE, _XL_REV_CODIGO = 0, 1, 2
_XL_REV_TIPO_SOCIEDAD, _XL_REV_CATEGORIA = 3, 4
_XL_REV_BENEF_DESDE, _XL_REV_BENEF_HASTA, _XL_REV_BENEF_TIPO = 9, 10, 11


def es_excel(ruta: str | Path) -> bool:
    return Path(ruta).suffix.lower() in _EXTENSIONES_EXCEL


def _filas_excel(ruta: Path) -> list[tuple]:
    try:
        import openpyxl
    except ImportError as exc:
        raise ErrorFuente(
            "Leer los reportes de SICAM en planilla necesita openpyxl:\n"
            "    pip install openpyxl"
        ) from exc

    try:
        libro = openpyxl.load_workbook(str(ruta), read_only=True, data_only=True)
    except Exception as exc:
        raise ErrorFuente(f"No pude abrir {ruta.name}: {exc}") from exc
    try:
        return [tuple(fila) for fila in libro[libro.sheetnames[0]].iter_rows(values_only=True)]
    finally:
        libro.close()


def _celda_excel(fila: tuple, indice: int) -> str:
    """Texto de una celda, sin los espacios duros que mete el export."""
    if indice >= len(fila) or fila[indice] is None:
        return ""
    return str(fila[indice]).replace("\xa0", " ").strip()


def _periodo_excel(fila: tuple, indice: int) -> Periodo | None:
    """El export trae los períodos como fecha o como ``MM/AAAA``."""
    if indice >= len(fila):
        return None
    valor = fila[indice]
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return Periodo(valor.year, valor.month)
    return _periodo(str(valor).replace("\xa0", " "))


def _monto_excel(fila: tuple, indice: int) -> Decimal:
    """Importe del export: formato anglosajón y espacios duros."""
    if indice >= len(fila) or fila[indice] is None:
        return CERO
    valor = fila[indice]
    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor))
    texto = str(valor).replace("\xa0", "").replace(" ", "").replace(",", "")
    if not texto or texto in {"-", "."}:
        return CERO
    try:
        return Decimal(texto)
    except InvalidOperation:
        return CERO


def leer_deuda_excel(ruta: str | Path) -> LecturaDeuda:
    """Lee el «Detalle de la Deuda» exportado a planilla."""
    ruta = Path(ruta)
    if not ruta.exists():
        raise ErrorFuente(f"No encontré el archivo: {ruta}")

    lectura = LecturaDeuda()
    descartadas = 0
    for fila in _filas_excel(ruta):
        desde = _periodo_excel(fila, _XL_DEUDA_DESDE)
        if desde is None:
            continue                      # encabezados y filas de título
        hasta = _periodo_excel(fila, _XL_DEUDA_HASTA) or desde
        if hasta < desde:
            descartadas += 1
            continue
        lectura.filas.append(
            FilaDeuda(
                desde=desde,
                hasta=hasta,
                categoria_historica=_celda_excel(fila, _XL_DEUDA_CAT_HISTORICA),
                categoria_actual=_celda_excel(fila, _XL_DEUDA_CAT_ACTUAL),
                beneficio_aplicado=_celda_excel(fila, _XL_DEUDA_BENEFICIO),
                meses_aportes=_monto_excel(fila, _XL_DEUDA_MESES_APORTES),
                capital_subtotal=_monto_excel(fila, _XL_DEUDA_CAPITAL_SUBTOTAL),
                intereses_subtotal=_monto_excel(fila, _XL_DEUDA_INTERESES_SUBTOTAL),
                total=_monto_excel(fila, _XL_DEUDA_TOTAL),
            )
        )

    if not lectura.filas:
        raise ErrorFuente(
            f"No reconocí ningún renglón en {ruta.name}. ¿Es el «Detalle de la "
            "Deuda» de SICAM exportado a planilla?"
        )
    if descartadas:
        lectura.advertencias.append(
            f"{descartadas} renglón(es) con período invertido; se omitieron."
        )
    return lectura


def leer_revista_excel(ruta: str | Path) -> LecturaRevista:
    """Lee la «Situación de Revista» exportada a planilla."""
    ruta = Path(ruta)
    if not ruta.exists():
        raise ErrorFuente(f"No encontré el archivo: {ruta}")

    lectura = LecturaRevista()
    for fila in _filas_excel(ruta):
        inicio = _periodo_excel(fila, _XL_REV_INICIO)
        if inicio is None:
            continue
        lectura.periodos.append(
            PeriodoRevista(
                inicio=inicio,
                cese=_periodo_excel(fila, _XL_REV_CESE),
                codigo_actividad=_celda_excel(fila, _XL_REV_CODIGO),
                tipo_sociedad=_celda_excel(fila, _XL_REV_TIPO_SOCIEDAD),
                categoria_optativa=_celda_excel(fila, _XL_REV_CATEGORIA),
                beneficio_desde=_periodo_excel(fila, _XL_REV_BENEF_DESDE),
                beneficio_hasta=_periodo_excel(fila, _XL_REV_BENEF_HASTA),
                tipo_beneficio=_celda_excel(fila, _XL_REV_BENEF_TIPO),
            )
        )

    if not lectura.periodos:
        raise ErrorFuente(
            f"No reconocí ningún tramo en {ruta.name}. ¿Es la «Situación de "
            "Revista» de SICAM exportada a planilla?"
        )
    return lectura


def leer_sicam(
    ruta_revista: str | Path,
    ruta_deuda: str | Path | None,
    cuil: str,
    nombre: str | None = None,
    politica: PoliticaDeuda | None = None,
    dpi: int = OCR_DPI,
):
    """Lee los dos PDF de SICAM y devuelve la historia laboral de autónomos."""
    revista = leer_revista(ruta_revista, dpi)
    deuda = leer_deuda(ruta_deuda, dpi) if ruta_deuda else None
    nombres = [Path(ruta_revista).name]
    if ruta_deuda:
        nombres.append(Path(ruta_deuda).name)
    return historia_desde_sicam(
        revista, deuda, cuil, nombre, politica, fuente=f"sicam:{'+'.join(nombres)}"
    )
