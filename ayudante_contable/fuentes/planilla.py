"""Importación de historia laboral desde planillas (CSV, TSV, XLSX).

Es la ruta recomendada para producción: el profesional descarga la historia
laboral desde Mi ANSES y la importa acá. No depende de que el portal mantenga
su HTML, no toca credenciales y deja el archivo original como respaldo del
informe.

El mapeo de columnas es tolerante: acepta los nombres habituales en castellano,
con o sin tildes, en cualquier orden.
"""

from __future__ import annotations

import csv
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Sequence

from ..modelo.dominio import (
    EstadoIngreso,
    HistoriaLaboral,
    Periodo,
    RegistroMensual,
    TipoAporte,
)
from .base import ErrorFuente

__all__ = ["ImportadorPlanilla", "leer_planilla", "ALIAS_COLUMNAS"]

_TILDES = str.maketrans("áéíóúÁÉÍÓÚüÜñÑ", "aeiouAEIOUuUnN")
_RE_NO_ALFANUM = re.compile(r"[^a-z0-9]+")

ALIAS_COLUMNAS: dict[str, tuple[str, ...]] = {
    "periodo": (
        "periodo", "periodo devengado", "mes", "mes ano", "fecha", "per",
        "periodo declarado", "mes y ano",
    ),
    "cuit_empleador": (
        "cuit", "cuit empleador", "cuit del empleador", "cuit cuil empleador",
        "cuit patronal", "cuit contribuyente",
    ),
    "empleador": (
        "empleador", "razon social", "empresa", "denominacion", "nombre empleador",
        "razon social empleador",
    ),
    "tipo": ("tipo", "tipo aporte", "regimen", "modalidad", "situacion de revista", "condicion"),
    "remuneracion_imponible": (
        "remuneracion", "remuneracion imponible", "remuneracion imponible 1",
        "base imponible", "sueldo", "sueldo bruto", "bruto", "rem imponible",
        "remuneracion bruta", "renta imponible", "remuneracion sipa",
    ),
    "aporte_declarado": (
        "aporte", "aporte declarado", "aportes declarados", "aporte sipa",
        "aporte personal", "aportes",
    ),
    "aporte_ingresado": (
        "aporte ingresado", "aportes ingresados", "monto ingresado", "ingresado",
        "importe ingresado", "pagado", "aporte efectivo",
    ),
    "estado_ingreso": (
        "estado", "estado ingreso", "estado del aporte", "situacion aporte",
        "ingreso", "pago", "cancelado", "aporte ingresado sn", "deuda",
    ),
    "observaciones": ("observaciones", "obs", "detalle", "nota", "notas", "comentario"),
}


def _clave(texto: str) -> str:
    limpio = _RE_NO_ALFANUM.sub(" ", texto.translate(_TILDES).lower()).strip()
    return re.sub(r"\s+", " ", limpio)


def _mapear_encabezados(encabezados: Sequence[str]) -> dict[str, int]:
    """Devuelve ``campo -> índice de columna`` a partir de los encabezados."""
    normalizados = [_clave(str(h or "")) for h in encabezados]
    mapa: dict[str, int] = {}

    for campo, alias in ALIAS_COLUMNAS.items():
        # Coincidencia exacta primero; si no, la primera que empiece igual.
        for indice, encabezado in enumerate(normalizados):
            if encabezado in alias and indice not in mapa.values():
                mapa[campo] = indice
                break
        else:
            for indice, encabezado in enumerate(normalizados):
                if not encabezado or indice in mapa.values():
                    continue
                if any(encabezado.startswith(a) or a.startswith(encabezado) for a in alias):
                    mapa[campo] = indice
                    break

    if "periodo" not in mapa:
        raise ErrorFuente(
            "La planilla no tiene una columna de período reconocible. "
            f"Encabezados leídos: {list(encabezados)}"
        )
    return mapa


_RE_MONEDA = re.compile(r"[^\d,.\-]")


def parsear_decimal(valor: Any) -> Decimal | None:
    """Interpreta importes en formato argentino (1.234,56) o anglosajón (1,234.56)."""
    if valor is None:
        return None
    if isinstance(valor, Decimal):
        return valor
    if isinstance(valor, (int, float)):
        return Decimal(str(valor))

    texto = _RE_MONEDA.sub("", str(valor).strip())
    if not texto or texto in {"-", ",", "."}:
        return None

    tiene_coma, tiene_punto = "," in texto, "." in texto
    if tiene_coma and tiene_punto:
        # El separador decimal es el que aparece último.
        decimal_es_coma = texto.rfind(",") > texto.rfind(".")
        texto = (
            texto.replace(".", "").replace(",", ".")
            if decimal_es_coma
            else texto.replace(",", "")
        )
    elif tiene_coma:
        entero, _, resto = texto.rpartition(",")
        # "1,234" con tres decimales es separador de miles, no decimal.
        texto = f"{entero}.{resto}" if len(resto) != 3 else texto.replace(",", "")
    elif tiene_punto and len(texto.rpartition(".")[2]) == 3 and texto.count(".") >= 1:
        # "1.234" es mil doscientos treinta y cuatro en formato local.
        entero, _, resto = texto.rpartition(".")
        if entero and entero.lstrip("-").isdigit():
            texto = texto.replace(".", "")

    try:
        return Decimal(texto)
    except InvalidOperation:
        return None


def _texto(fila: Sequence[Any], mapa: dict[str, int], campo: str) -> str | None:
    indice = mapa.get(campo)
    if indice is None or indice >= len(fila):
        return None
    valor = fila[indice]
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto or None


def _fila_a_registro(
    fila: Sequence[Any], mapa: dict[str, int], numero: int, advertencias: list[str]
) -> RegistroMensual | None:
    crudo_periodo = _texto(fila, mapa, "periodo")
    if not crudo_periodo:
        return None
    try:
        periodo = Periodo.desde_texto(crudo_periodo)
    except (ValueError, TypeError):
        advertencias.append(f"Fila {numero}: período ilegible ({crudo_periodo!r}); se omitió.")
        return None

    remuneracion = parsear_decimal(_texto(fila, mapa, "remuneracion_imponible")) or Decimal("0")
    cuit = _texto(fila, mapa, "cuit_empleador")
    if cuit:
        cuit = re.sub(r"\D", "", cuit) or None

    return RegistroMensual(
        periodo=periodo,
        cuit_empleador=cuit,
        empleador=_texto(fila, mapa, "empleador"),
        tipo=TipoAporte.desde_texto(_texto(fila, mapa, "tipo")),
        remuneracion_imponible=remuneracion,
        aporte_declarado=parsear_decimal(_texto(fila, mapa, "aporte_declarado")),
        aporte_ingresado=parsear_decimal(_texto(fila, mapa, "aporte_ingresado")),
        estado_ingreso=EstadoIngreso.desde_texto(_texto(fila, mapa, "estado_ingreso")),
        observaciones=_texto(fila, mapa, "observaciones") or "",
    )


def _filas_csv(ruta: Path) -> list[list[str]]:
    for codificacion in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            contenido = ruta.read_text(encoding=codificacion)
            break
        except UnicodeDecodeError:
            continue
    else:  # pragma: no cover - latin-1 nunca falla
        raise ErrorFuente(f"No pude decodificar {ruta}")

    muestra = contenido[:4096]
    try:
        dialecto = csv.Sniffer().sniff(muestra, delimiters=";,\t|")
        delimitador = dialecto.delimiter
    except csv.Error:
        delimitador = ";" if muestra.count(";") > muestra.count(",") else ","

    return [fila for fila in csv.reader(contenido.splitlines(), delimiter=delimitador) if fila]


def _filas_xlsx(ruta: Path) -> list[list[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depende del entorno
        raise ErrorFuente(
            "Para leer archivos .xlsx instalá openpyxl:  pip install openpyxl\n"
            "También podés exportar la planilla a CSV y usar ese archivo."
        ) from exc

    libro = load_workbook(ruta, read_only=True, data_only=True)
    hoja = libro.active
    filas = [list(fila) for fila in hoja.iter_rows(values_only=True)]
    libro.close()
    return [f for f in filas if any(c is not None and str(c).strip() for c in f)]


def _buscar_encabezado(filas: list[Sequence[Any]]) -> int:
    """Ubica la fila de encabezados, salteando títulos previos del reporte."""
    for indice, fila in enumerate(filas[:20]):
        normalizados = {_clave(str(c or "")) for c in fila}
        if normalizados & set(ALIAS_COLUMNAS["periodo"]):
            return indice
        if any(
            any(n.startswith(a) for a in ALIAS_COLUMNAS["periodo"]) for n in normalizados if n
        ):
            return indice
    return 0


def leer_planilla(ruta: str | Path, cuil: str, nombre: str | None = None) -> HistoriaLaboral:
    """Lee una planilla de historia laboral y la convierte al modelo interno."""
    ruta = Path(ruta)
    if not ruta.exists():
        raise ErrorFuente(f"No encontré el archivo: {ruta}")

    if ruta.suffix.lower() in {".xlsx", ".xlsm"}:
        filas = _filas_xlsx(ruta)
    else:
        filas = _filas_csv(ruta)

    if not filas:
        raise ErrorFuente(f"El archivo {ruta.name} está vacío.")

    indice_encabezado = _buscar_encabezado(filas)
    mapa = _mapear_encabezados(filas[indice_encabezado])

    advertencias: list[str] = []
    registros: list[RegistroMensual] = []
    for desplazamiento, fila in enumerate(filas[indice_encabezado + 1 :], start=1):
        registro = _fila_a_registro(
            fila, mapa, indice_encabezado + 1 + desplazamiento, advertencias
        )
        if registro is not None:
            registros.append(registro)

    if not registros:
        raise ErrorFuente(
            f"No pude extraer ningún período de {ruta.name}. "
            "Revisá que la columna de período tenga formato MM/AAAA o AAAAMM."
        )

    faltantes = [c for c in ("aporte_ingresado", "estado_ingreso") if c not in mapa]
    if len(faltantes) == 2:
        advertencias.append(
            "La planilla no trae columna de aporte ingresado ni de estado: el control "
            "de ingreso efectivo queda como 'sin dato' en todos los períodos."
        )

    return HistoriaLaboral(
        cuil=cuil,
        registros=registros,
        nombre=nombre,
        fuente=f"planilla:{ruta.name}",
        advertencias_origen=advertencias,
    )


class ImportadorPlanilla:
    """Adaptador de :class:`FuenteHistoriaLaboral` sobre un archivo local."""

    nombre = "planilla"

    def __init__(self, ruta: str | Path, nombre_afiliado: str | None = None) -> None:
        self.ruta = Path(ruta)
        self.nombre_afiliado = nombre_afiliado

    def obtener(self, cuil: str) -> HistoriaLaboral:
        return leer_planilla(self.ruta, cuil, self.nombre_afiliado)
